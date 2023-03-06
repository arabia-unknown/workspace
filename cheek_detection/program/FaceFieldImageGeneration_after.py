import dlib
import cv2
import numpy as np
import openpyxl
import glob
from natsort import natsorted
from imutils import face_utils


#顔検出ツールの呼び出し
detector = dlib.get_frontal_face_detector()
#顔のランドマーク検出ツールの呼び出し
predictor = dlib.shape_predictor("shape_predictor_68_face_landmarks.dat")
#空のExcelファイルを作成
wb1 = openpyxl.Workbook()
wb2 = openpyxl.Workbook()
wb3 = openpyxl.Workbook()
wb4 = openpyxl.Workbook()
wb5 = openpyxl.Workbook()
#ワークシートを作成
ws1 = wb1.active
ws2 = wb2.active
ws3 = wb3.active
ws4 = wb4.active
ws5 = wb5.active

Width = 160
Height = 0

#画像番号から顔領域を算出
def FaceFieldDetection(number):
    xpoint = []
    ypoint = []
    for i in range(34):
        if (i % 2 == 0):
            xpoint.append(contoursAreaData[number, i])
        else:
            ypoint.append(contoursAreaData[number, i])
    for i in range(20):
        if (i % 2 == 0):
            xpoint.append(eyebrowsAreaData[number, i])
        else:
            ypoint.append(eyebrowsAreaData[number, i])
    x_max = max(xpoint)
    x_min = min(xpoint)
    y_max = max(ypoint)
    y_min = min(ypoint) - 30
    if (y_min <= 0):
        y_min = 0
    rect = [x_min, y_min, x_max - x_min, y_max - y_min]
    return rect
    


#Excelファイルから眉の領域の座標データを読み込む
#眉の座標データが書かれているExcelデータまでのパスを入力してください
wb_r2 = openpyxl.load_workbook(r"C:\Users\horii\Desktop\sagyou\program\elderly_autocheek\A\10\practice2\after\excel\Eyebrows.xlsx")
ws_r2 = wb_r2["Sheet"]
max_row = ws_r2.max_row
data2 = []
for r in range(max_row):
    column = 0
    row = ws_r2[r+1]
    while column < 20 :
        data2.append(row[column].value)
        column += 1
#被験者の眉領域データ
eyebrowsAreaData = np.array(data2).reshape(max_row,20)

#Excelファイルから輪郭の領域の座標データを読み込む
#輪郭の座標データが書かれているExcelデータまでのパスを入力してください
wb_r3 = openpyxl.load_workbook(r"C:\Users\horii\Desktop\sagyou\program\elderly_autocheek\A\10\practice2\after\excel\Contours.xlsx")
ws_r3 = wb_r3["Sheet"]
max_row = ws_r3.max_row
data3 = []
for r in range(max_row):
    column = 0
    row = ws_r3[r+1]
    while column < 34 :
        data3.append(row[column].value)
        column += 1
#被験者の頬領域データ
contoursAreaData = np.array(data3).reshape(max_row,34)

#Excelファイルから鼻の領域の座標データを読み込む
#鼻の座標データが書かれているExcelデータまでのパスを入力してください
wb_r4 = openpyxl.load_workbook(r"C:\Users\horii\Desktop\sagyou\program\elderly_autocheek\A\10\practice2\after\excel\nose.xlsx")
ws_r4 = wb_r4["Sheet"]
max_row = ws_r4.max_row
data4 = []
for r in range(max_row):
    column = 0
    row = ws_r4[r+1]
    while column < 18 :
        data4.append(row[column].value)
        column += 1
#被験者の頬領域データ
noseAreaData = np.array(data4).reshape(max_row,18)

#Excelファイルから頬の領域の座標データを読み込む
#頬の座標データが書かれているExcelデータまでのパスを入力してください
wb_r5 = openpyxl.load_workbook(r"C:\Users\horii\Desktop\sagyou\program\elderly_autocheek\A\10\practice2\after\excel\cheek.xlsx")
ws_r5 = wb_r5["Sheet"]
max_row = ws_r5.max_row
data5 = []
for r in range(max_row):
    column = 0
    row = ws_r5[r+1]
    while column < 8 :
        data5.append(row[column].value)
        column += 1
#被験者の頬領域データ
cheekAreaData = np.array(data5).reshape(max_row,8)




#画像ファイルを取得
files = []
#対象とする画像が入っているファイルまでのパスを入力してください
for f_name in glob.iglob(r"F:\ryouta\Program\Experiment\1019\A\Practice2\Image\after\original\*.jpg"):
    files.append(f_name)
files = natsorted(files)

img_num = len(files)
for i in range(img_num):
    img = cv2.imread(files[i])
    #顔領域を検出
    faceField = FaceFieldDetection(i)
    
    #画像から顔領域を切り取る
    startPoint_Left = faceField[0]
    startPoint_Top = faceField[1]
    startPoint_Right = startPoint_Left + faceField[2]
    startPoint_Bottom = startPoint_Top + faceField[3]
    cut_img = img[startPoint_Top : startPoint_Bottom, startPoint_Left : startPoint_Right] #画像から指定領域を切り出し
    print(startPoint_Left,startPoint_Top)

    #画像の出力先を入力してください
    #cv2.imwrite(r"D:\ryouta\Program\1014_Pre_experiment\honda\ex\landmark\cut\cut_img_" + str(i) + "_.jpg",cut_img)    
    cv2.imwrite(r"C:\Users\horii\Desktop\sagyou\program\elderly_autocheek\A\10\practice2\after\cut\cut_img_" + str(i) + "_.jpg",cut_img)
    
    
    #切り取った顔画像をリサイズし出力(ゲームプレイ前の場合)
    # if (i == 0):
    #     #切り取った顔画像をリサイズ
    #     #幅を160にするには、現在の幅の何倍かを計算(scale)
    #     #高さも横幅と同じだけの倍率増やす(Height * scale)
    #     x_scale = 160 / float(faceField[2])
    #     resize_Height = int(round(faceField[3] * x_scale))
    #     re_img = cv2.resize(cut_img, dsize=(Width, resize_Height))
    #     Height = resize_Height
    #     y_scale = Height / float(faceField[3])
    
    #     #処理後の画像を出力
    #     #画像を出力するファイルまでのパスを入力してください         
    #     #cv2.imwrite(r"D:\ryouta\Program\1014_Pre_experiment\honda\ex\landmark\result\re_img_" + str(i) + "_.jpg",re_img)
    #     cv2.imwrite(r"C:\Users\horii\Desktop\sagyou\program\elderly_autocheek\A\10\practice1\after\re\re_img_" + str(i) + "_.jpg",re_img)


    # # #切り取った顔画像をリサイズ
    # # #幅を160にするには、現在の幅の何倍かを計算(scale)
    # # #高さも横幅と同じだけの倍率増やす(Height * scale)
    # # x_scale = 160 / float(faceField[2])
    # # resize_Height = int(round(faceField[3] * x_scale))
    # # re_img = cv2.resize(cut_img, dsize=(Width, resize_Height))
    # # Height = resize_Height
    # # y_scale = Height / float(faceField[3])
    
    # # #処理後の画像を出力
    # # #画像を出力するファイルまでのパスを入力してください         
    # # #cv2.imwrite(r"D:\ryouta\Program\1014_Pre_experiment\honda\ex\landmark\result\re_img_" + str(i) + "_.jpg",re_img)
    # # cv2.imwrite(r"C:\Users\horii\Desktop\sagyou\program\re\re_img_" + str(i) + "_.jpg",re_img)


    
    # else:
    #     #切り取った顔画像をリサイズ
    #     re_img = cv2.resize(cut_img, dsize=(Width, Height))
    
    #     #処理後の画像を出力
    #     #画像を出力するファイルまでのパスを入力してください        
    #     #cv2.imwrite(r"D:\ryouta\Program\1014_Pre_experiment\honda\ex\landmark\result\re_img_" + str(i) + "_.jpg",re_img)
    #     cv2.imwrite(r"C:\Users\horii\Desktop\sagyou\program\elderly_autocheek\A\10\practice1\after\re\re_img_" + str(i) + "_.jpg",re_img)
    

    #切り取った顔画像をリサイズ(ゲームプレイ後の場合)
    #ゲームプレイ前におけるリサイズ画像と、同じサイズを入力してください(変更点は数値の部分のみ)
    x_scale = 160 / float(faceField[2])
    y_scale = 188 / float(faceField[3])
    re_img = cv2.resize(cut_img, dsize=(160, 188))
    
    #処理後の画像を出力
    #画像を出力するファイルまでのパスを入力してください         
    cv2.imwrite(r"C:\Users\horii\Desktop\sagyou\program\elderly_autocheek\A\10\practice2\after\re\re_img_" + str(i) + "_.jpg",re_img)

    
    #座標情報を更新    
    #眉の領域の座標データ
    y = 0
    while y < 20:
        value_x = int(round(float(eyebrowsAreaData[i, y] - startPoint_Left) * x_scale))
        ws2.cell(row=i+1, column=y+1, value=value_x)
        y += 1
        value_y = int(round(float(eyebrowsAreaData[i, y] - startPoint_Top) * y_scale))
        ws2.cell(row=i+1, column=y+1, value=value_y)
        y += 1
    #輪郭の領域の座標データ
    z = 0
    while z < 34:
        value_x = int(round(float(contoursAreaData[i, z] - startPoint_Left) * x_scale))
        ws3.cell(row=i+1, column=z+1, value=value_x)
        z += 1
        value_y = int(round(float(contoursAreaData[i, z] - startPoint_Top) * y_scale))
        ws3.cell(row=i+1, column=z+1, value=value_y)
        z += 1
    #鼻の領域の座標データ
    w = 0
    while w < 18:
        value_x = int(round(float(noseAreaData[i, w] - startPoint_Left) * x_scale))
        ws4.cell(row=i+1, column=w+1, value=value_x)
        w += 1
        value_y = int(round(float(noseAreaData[i, w] - startPoint_Top) * y_scale))
        ws4.cell(row=i+1, column=w+1, value=value_y)
        w += 1

    #頬の領域の座標データ
    v = 0
    print(x_scale,y_scale)
    while v < 8:
        value_x = int(round(float(cheekAreaData[i, v] - startPoint_Left) * x_scale))
        ws5.cell(row=i+1, column=v+1, value=value_x)
        v += 1
        value_y = int(round(float(cheekAreaData[i, v] - startPoint_Top) * y_scale))
        ws5.cell(row=i+1, column=v+1, value=value_y)
        v += 1

#座標データを出力
#Excelデータを出力するファイルまでのパスを入力してください。
#wb2.save(r"D:\ryouta\Program\1014_Pre_experiment\honda\ex\landmark\outfile\Re_CoordinatesData(Eyebrows).xlsx")
wb2.save(r"C:\Users\horii\Desktop\sagyou\program\elderly_autocheek\A\10\practice2\after\excel\re_Eyebrows.xlsx")
#wb3.save(r"D:\ryouta\Program\1014_Pre_experiment\honda\ex\landmark\outfile\Re_CoordinatesData(Contours).xlsx")
wb3.save(r"C:\Users\horii\Desktop\sagyou\program\elderly_autocheek\A\10\practice2\after\excel\re_Contours.xlsx")
#wb4.save(r"D:\ryouta\Program\1014_Pre_experiment\honda\ex\landmark\outfile\Re_CoordinatesData(nose).xlsx")
wb4.save(r"C:\Users\horii\Desktop\sagyou\program\elderly_autocheek\A\10\practice2\after\excel\re_nose.xlsx")
wb5.save(r"C:\Users\horii\Desktop\sagyou\program\elderly_autocheek\A\10\practice2\after\excel\re_cheek.xlsx")
   
        
    
    
    












