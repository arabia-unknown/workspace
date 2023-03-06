import cv2
import numpy as np
import openpyxl
import glob
from natsort import natsorted
import os
import statistics
import math
import gc


#L*a*b色空間の定義
ColorLab_All = []
ColorLab_Cheek = []
#ＸＹＺ色空間の定義
ColorXyz_All = []
ColorXyz_Cheek = []
#HSV色空間の定義
ColorHsv = []

flag = 0

#Excelファイルを読み込む
#相対彩度割合の情報を書き込むExcelファイルまでのパスを入力してください。
#colorDataWorkbookSaturation_Lab = openpyxl.load_workbook(r"D:\ryouta\Program\Experiment\1124\M\Left_sRGB-D50_LabSaturationColorDataCheek(normalization,Dlib).xlsx")
colorDataWorkbookSaturation_Lab = openpyxl.load_workbook(r"C:\Users\horii\Desktop\sagyou\program\elderly_autocheek\A\10\practice2\after\excel\Left_sRGB-D50_LabSaturationColorDataCheek(normalization,Dlib)_Ya.xlsx")
#対象のワークシートを読み込む
colorDataWorksheetSaturation_Lab = colorDataWorkbookSaturation_Lab['Sheet1']


#顔領域全体の彩度値を保存する配列
saturationImageData_All = []
#頬領域の彩度値を保存する配列
saturationImageData_Cheek = []
#頬領域の彩度値が顔領域全体の何番目に属しているかを保存する配列
cheekSaturationPositioning = []
#部分処理を行う矩形データ(rect = [X座標, Y座標, Width, Height])(必要に応じて変更)
rect = [0,0,30,30]
#被験者の頬領域データ
cheekAreaData = []
#目領域を検出したかどうかの確認
detectionFlag = True


#Excelファイルから頬の相対座標データを読み込む
#頬の相対座標データが書かれているExcelデータまでのパスを入力してください
# wb_r1 = openpyxl.load_workbook(r"D:\ryouta\Program\Experiment\1124\M\Match\Image\after\outfile\match_after_RelativeCoordinatesData.xlsx")
wb_r1 = openpyxl.load_workbook(r"C:\Users\horii\Desktop\sagyou\program\elderly_autocheek\A\10\practice2\after\excel\RelativeCoordinatesData.xlsx")
ws_r1 = wb_r1["Sheet"]
#読み込んだデータをリストに入力
cheekAreaData.append(ws_r1["A2"].value)
cheekAreaData.append(ws_r1["B2"].value)
cheekAreaData.append(ws_r1["C2"].value)
cheekAreaData.append(ws_r1["D2"].value)
cheekAreaData.append(ws_r1["E2"].value)



#頬領域の抽出に鼻領域を使用する場合
#Excelファイルから鼻の領域の座標データを読み込む
#鼻の座標データが書かれているExcelデータまでのパスを入力してください
# wb_r2 = openpyxl.load_workbook(r"D:\ryouta\Program\Experiment\1124\M\Match\Image\after\outfile\Re_CoordinatesData(nose).xlsx")
wb_r2 = openpyxl.load_workbook(r"C:\Users\horii\Desktop\sagyou\program\elderly_autocheek\A\10\practice2\after\excel\re_nose.xlsx")
ws_r2 = wb_r2["Sheet"]
max_row = ws_r2.max_row
data = []
for r in range(max_row):
    column = 0
    row = ws_r2[r+1]
    while column < 18 :
        data.append(row[column].value)
        column += 1
#被験者の鼻領域データ
noseAreaData = np.array(data).reshape(max_row,18)



#処理対象領域を選択
Target = input("対象領域を選択して下さい (right or left) :")


        
#画像ファイルを取得
#対象とする画像までのパスを入力してください
files = []
for f_name in glob.iglob(r"C:\Users\horii\Desktop\sagyou\program\elderly_autocheek\A\10\practice2\after\re\*.jpg"):
    files.append(f_name)
files = natsorted(files)



#出力配列の定義
outputData = [[0] * len(files) for i in range(12)]
#被験者名の設定
testerName = os.path.basename(files[0]).split('_', 4)[0]



#処理対象の頬領域の相対座標を取得
index = 0
cheek_len = len(cheekAreaData)
for j in range(cheek_len):
    if (cheekAreaData[j] == testerName):
        index = j
if (Target == "right"):
    rightCheek_X = cheekAreaData[index + 1]
    rightCheek_Y = cheekAreaData[index + 2]
else:
    leftCheek_X = cheekAreaData[index + 3]
    leftCheek_Y = cheekAreaData[index + 4]



#RGB色空間からXYZ色空間への変換
def ChangeRgbtoXyz(B,G,R):
    #元画像のRGB値
    r = R
    g = G
    b = B
    #色空間への変換(sRGBからXYZ(D50光源))
    X = 0.4124000 * (r / 255) * 100 + 0.3576000 * (g / 255) * 100 + 0.1805000 * (b / 255) * 100
    Y = 0.2126000 * (r / 255) * 100 + 0.7152000 * (g / 255) * 100 + 0.0722000 * (b / 255) * 100
    Z = 0.0193000 * (r / 255) * 100 + 0.1192000 * (g / 255) * 100 + 0.9505000 * (b / 255) * 100
    returnXyz = [X,Y,Z]
    return returnXyz

    
#XYZ色空間からLab色空間への変換
def ChangeXyztoLab(X,Y,Z):
    #元画像のXYZ値
    imageXYZ_X = X
    imageXYZ_Y = Y
    imageXYZ_Z = Z
    #三刺激値
    TristimulusValueX = 98.072
    TristimulusValueY = 100.000
    TristimulusValueZ = 118.225
    #L*の設定 L*の範囲
    L1 = imageXYZ_Y / TristimulusValueY
    L2 = math.pow((imageXYZ_Y / TristimulusValueY), 1/3)
    L3 = imageXYZ_Y / TristimulusValueY
    if (L1 < 0.008856):
        LStar = (116 * L2) - 16
    else:
        LStar = 903.29 * L3
    #a*の設定 a*の範囲
    a1 = math.pow((imageXYZ_X / TristimulusValueX), 1/3)
    a2 = math.pow((imageXYZ_Y / TristimulusValueY), 1/3)
    aStar = 504.3 * (a1 - a2)
    #b*の設定 b*の範囲
    b1 = math.pow((imageXYZ_Y / TristimulusValueY), 1/3)
    b2 = math.pow((imageXYZ_Z / TristimulusValueZ), 1/3)
    bStar = 201.7 * (b1 - b2)
    #habの取得(math.atan2=xとyのアークタンジェント座標値を返す)
    hab = math.atan2(bStar, aStar) * (180 / math.pi)
    returnLab = [LStar,aStar,bStar,hab]
    return returnLab

    
#Lab色空間においての彩度の算出
def SaturationCalculation_Lab(aStar,bStar):
    S = math.sqrt(math.pow(aStar,2) + math.pow(bStar, 2))
    return S


#サーモグラフィ的色変換
def ColorScaleBCGYR(in_value):
    #0.0~1.0の範囲の値をサーモグラフィみたいな色にする
    #0.0              1.0
    #青　　水　　緑　　黄　　赤
    #最小値以下 = 青
    #最大値以上 = 赤
    a = 255
    value = in_value
    tmp_val = math.cos(4 * math.pi * value)
    col_val = (int)((-tmp_val / 2 + 0.5) * 255)
    #赤
    if (value >= (4.0/4.0)):
        r = 255
        g = 0
        b = 0
    #黄～赤
    elif (value >= (3.0/4.0)):
        r = 255
        g = col_val
        b = 0
    #緑～黄
    elif (value >= (2.0/4.0)):
        r = col_val
        g = 255
        b = 0
    #水～緑
    elif (value >= (1.0/4.0)):
        r = 0
        g = 255
        b = col_val
    #青～水
    elif (value >= (0.0/4.0)):
        r = 0
        g = col_val
        b = 255
    #青
    else:
        r = 0
        g = 0
        b = 255
    #ret = (a & 0x000000FF) << 24 | (r & 0x000000FF) << 16 | (g & 0x000000FF) << 8 | (b & 0x000000FF)
    ret = [a, b, g, r]
    return ret    



#色情報の読み込み
def ColorSet(fileName):
    #画像ファイルの読み込み
    img = cv2.imread(fileName)
    sa_All_img = cv2.imread(fileName)
    sa_cheek_img = cv2.imread(fileName)
    #対象領域のセット(画像の上部15%をカットし、矩形に外接する楕円の外側をカットした領域を指定)
    targetreArea = []
    height, width, channels = img.shape[:3]
    for y in range(height):
        for x in range(width):
            w = width
            h = height
            mx = x - ((w-1) / 2)
            my = y - ((h-1) / 2)
            a = w / 2
            b = h / 2
            if ((pow(mx, 2) / pow(a, 2)) + (pow(my, 2) / pow(b, 2)) > 1 or (y <= 0.15 * h)):
                pass
            else:
                targetreArea.append(x)
                targetreArea.append(y)
    target_val = int(len(targetreArea) / 2)
    targetreArea = np.array(targetreArea).reshape(target_val, 2)

  
    
    if (detectionFlag):
        saturationImageData_All.clear()
        saturationImageData_Cheek.clear()
        cheekSaturationPositioning.clear()
        ColorXyz_All.clear()
        ColorLab_All.clear()
        ColorXyz_Cheek.clear()
        ColorLab_Cheek.clear()        
        #全画素(targetArea内)の彩度値を取得
        k = 0
        while k < target_val:
            p_X = targetreArea[k][0]
            p_Y = targetreArea[k][1]
            #それぞれの画素値を取得
            blue = img[p_Y, p_X, 0]
            green = img[p_Y, p_X, 1]
            red = img[p_Y, p_X, 2]
            ColorXyz_All.append(ChangeRgbtoXyz(blue,green,red))
            Color_X = ColorXyz_All[k][0]
            Color_Y = ColorXyz_All[k][1]
            Color_Z = ColorXyz_All[k][2]
            ColorLab_All.append(ChangeXyztoLab(Color_X, Color_Y, Color_Z))
            Color_a = ColorLab_All[k][1]
            Color_b = ColorLab_All[k][2]
            saturationImageData_All.append(SaturationCalculation_Lab(Color_a, Color_b))
            """
            #全画素(targetArea内)の彩度画像を作成する場合
            min = 12.5
            max = 21.0
            afterColor = ColorScaleBCGYR((saturationImageData_All[k] - min) / (max - min))
            sa_All_img[p_Y, p_X] = afterColor[1], afterColor[2], afterColor[3]            
            """
            k += 1
        
        #全画素(targetArea内)の彩度画像を出力
        #cv2.imwrite(r"D:\ryouta\Program\Experiment\1019\A\Practice1\Image\before\check\All_saturation_img_" + str(flag) + ".jpg",sa_All_img)
        
        
        #矩形情報を取得(鼻と頬の相対座標から矩形の位置を決める)
        if (Target == "right"):
            rect[0] = nose_right_X + rightCheek_X
            rect[1] = nose_right_Y + rightCheek_Y
            
        else:
            rect[0] = nose_left_X + leftCheek_X
            rect[1] = nose_left_Y + leftCheek_Y
            
        
        #頬領域の彩度取得
        y = rect[1]
        rect_Bottom = y + rect[3]
        l = 0
        while (y < rect_Bottom):
            x = rect[0]
            rect_Right = x + rect[2]
            while (x < rect_Right):
                w = width
                h = height
                mx = x - ((w - 1) / 2)
                my = y - ((h - 1) / 2)
                a = w / 2
                b = h / 2
                if ((math.pow(mx, 2) / math.pow(a, 2)) + (math.pow(my, 2) / math.pow(b, 2)) > 1 or (y <= 0.15 * h)):
                    pass
                else:
                    blue = img[y, x, 0]
                    green = img[y, x, 1]
                    red = img[y, x, 2]
                    ColorXyz_Cheek.append(ChangeRgbtoXyz(blue,green,red))
                    #ここでIndexErrorは、頬領域が画像からはみ出している時
                    Color_X = ColorXyz_Cheek[l][0]
                    Color_Y = ColorXyz_Cheek[l][1]
                    Color_Z = ColorXyz_Cheek[l][2]
                    ColorLab_Cheek.append(ChangeXyztoLab(Color_X, Color_Y, Color_Z))
                    Color_a = ColorLab_Cheek[l][1]
                    Color_b = ColorLab_Cheek[l][2]
                    saturationImageData_Cheek.append(SaturationCalculation_Lab(Color_a, Color_b))
                    
                    #頬領域の彩度画像を作成する場合
                    min = 10.0
                    max = 20.0
                    afterColor = ColorScaleBCGYR((saturationImageData_Cheek[l] - min) / (max - min))
                    sa_cheek_img[y, x] = afterColor[1], afterColor[2], afterColor[3]
                    
                l += 1
                x += 1
            y += 1
        
        
        #頬領域の彩度画像を出力
        #cv2.imwrite(r"D:\ryouta\Program\Experiment\1124\M\Match\Image\after\resize\left_cheek\Cheek_saturation_img_" + str(flag) + ".jpg",sa_cheek_img)
        
        
                   
        #データを昇順ソート
        saturationImageData_All.sort()
        saturationImageData_Cheek.sort()
        #"cheekSaturationPositioning"には、、顔領域内の何番目に頬領域の情報が入っているかを表す
        #つまり、頬領域の彩度情報(最小値や最大値)が、顔領域全体で見た時に、何番目のセルに存在しているのか
        m = -1
        for i in range(len(saturationImageData_Cheek)):
            positioningDetectionFlag = False
            while (positioningDetectionFlag == False):
                m += 1
                if (saturationImageData_Cheek[i] == saturationImageData_All[m]):
                    cheekSaturationPositioning.append(m)
                    positioningDetectionFlag = True
    else:
        saturationImageData_All.clear()
        saturationImageData_Cheek.clear()
        cheekSaturationPositioning.clear()
        saturationImageData_All.append(-1)
        saturationImageData_Cheek.append(-1)
        cheekSaturationPositioning.append(-1)
    
    


#画像から読み込んだ色情報をExcelに出力
#エクセル書き込み用のデータセット
img_data = len(files)
for i in range(img_data):
    #撮影タイミングの取得
    shootingTiming = os.path.basename(files[i]).split('_', 4)[3]
    """
    "os.path.basename()"は、拡張子を含むファイル名の文字列を返す
    "split"は文字列を指定した区切り文字で分割する。第1引数で区切り文字の指定を行い、第2引数で最大分割数を指定する。分割した文字列は[0],[1]などで指定できる
    """
    
    nose_right_X = noseAreaData[i][8]
    nose_right_Y = noseAreaData[i][9]
    nose_left_X = noseAreaData[i][16]
    nose_left_Y = noseAreaData[i][17]
    
    if (shootingTiming != os.path.basename(files[0]).split('_', 4)[3]):
        if (testerName != os.path.basename(files[0]).split('_', 4)[0]):
            #色情報をExcelに保存
            if (outputData[0,i] == -1):
                continue
            #データの入力
            for s in range(12):
                colorDataWorksheetSaturation_Lab.cell(row=((i % img_data)+5), column=s+3, value=outputData[s][i])
            
            #Excelファイルを保存
            colorDataWorkbookSaturation_Lab.save(r"C:\Users\horii\Desktop\sagyou\program\elderly_autocheek\A\10\practice2\after\excel\sRGB-D50_LabSaturationColorDataCheek(normalization,Dlib).xlsx")
            #被験者名の更新
            testerName = os.path.basename(files[0]).split('_', 4)[0]
            index = 0
            cheek_len = len(cheekAreaData)
            for j in range(cheek_len):
                if (cheekAreaData[j] == testerName):
                    index = j
                if (Target == "right"):
                    rightCheek_X = cheekAreaData[index + 1]
                    rightCheek_Y = cheekAreaData[index + 2]
                else:
                    leftCheek_X = cheekAreaData[index + 3]
                    leftCheek_Y = cheekAreaData[index + 4]
        shootingTiming = os.path.basename(files[0]).split('_', 4)[3]
        
    
    ColorSet(files[i])
    flag += 1
    col = i % img_data
    if (detectionFlag):
        outputData[0][col] = len(saturationImageData_All)                    #総画素数(対象領域全体)
        outputData[1][col] = len(saturationImageData_Cheek)                  #総画素数(頬領域)
        outputData[2][col] = max(saturationImageData_Cheek)                  #最大値(頬領域の彩度)
        outputData[3][col] = min(saturationImageData_Cheek)                  #最小値(頬領域の彩度)
        outputData[4][col] = statistics.mean(saturationImageData_Cheek)      #平均値(頬領域の彩度)
        outputData[5][col] = statistics.median(saturationImageData_Cheek)    #中央値(頬領域の彩度)
        outputData[6][col] = max(cheekSaturationPositioning)                 #最大値(顔領域全体における頬領域の画素の位置)
        outputData[7][col] = min(cheekSaturationPositioning)                 #最小値(顔領域全体における頬領域の画素の位置)
        outputData[8][col] = statistics.mean(cheekSaturationPositioning)     #平均値(顔領域全体における頬領域の画素の位置)
        outputData[9][col] = statistics.median(cheekSaturationPositioning)   #中央値(顔領域全体における頬領域の画素の位置)
        outputData[10][col] = outputData[8][col] / outputData[0][col]        #割合(顔領域全体における頬領域の画素の位置の平均)
        outputData[11][col] = outputData[9][col] / outputData[0][col]        #割合(顔領域全体における頬領域の画素の位置の中央)
    else:
        outputData[0][col] = -1
    #ガベレージコレクション
    gc.collect()

#色情報をExcelに保存
#c = 60
c = 36
for o in range(img_data):
    if (outputData[0][o] == -1):
        continue
    #データの入力(pにプラスする値を変更することで、練習何回目かを決定する)
    for p in range(12):
        colorDataWorksheetSaturation_Lab.cell(row=((o % img_data)+5), column=p+3+c, value=outputData[p][o])

#Excelファイルを保存
#Excelファイルを出力するファイルまでのパスを入力してください
colorDataWorkbookSaturation_Lab.save(r"C:\Users\horii\Desktop\sagyou\program\elderly_autocheek\A\10\practice2\after\excel\Left_sRGB-D50_LabSaturationColorDataCheek(normalization,Dlib)_Ya.xlsx")