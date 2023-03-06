import dlib
import cv2
from imutils import face_utils
import numpy as np
import openpyxl
import glob
from natsort import natsorted

#顔検出ツールの呼び出し
detector = dlib.get_frontal_face_detector()
#顔のランドマーク検出ツールの呼び出し
predictor = dlib.shape_predictor("shape_predictor_68_face_landmarks.dat")

#空のExcelファイルを作成
wb = openpyxl.Workbook()
#ワークシートを作成
ws = wb.active

#Excelファイルからresize後の鼻領域の座標データを読み込む
#鼻の座標データが書かれているExcelデータまでのパスを入力してください
#wb_r4 = openpyxl.load_workbook(r"D:\ryouta\Program\1014_Pre_experiment\honda\ex\landmark\outfile\Re_CoordinatesData(nose).xlsx")
wb_r4 = openpyxl.load_workbook(r"C:\Users\horii\Desktop\sagyou\program\elderly_autocheek\A\10\practice2\after\excel\re_nose.xlsx")
ws_r4 = wb_r4["Sheet"]
max_row = ws_r4.max_row
data4 = []
for r in range(max_row):
    column = 0
    row = ws_r4[r+1]
    while column < 18 :
        data4.append(row[column].value)
        column += 1
 #被験者の鼻領域データ
noseAreaData = np.array(data4).reshape(max_row,18)

#頬領域を描画(手動で座標を入力)
wb_r5 = openpyxl.load_workbook(r"C:\Users\horii\Desktop\sagyou\program\elderly_autocheek\A\10\practice2\after\excel\re_cheek.xlsx")
ws_r5 = wb_r5["Sheet"]
max_row = ws_r5.max_row
data5 = []
for r in range(max_row):
    column = 0
    row = ws_r5[r+1]
    while column < 8 :
        data5.append(row[column].value)
        column += 1
#被験者の頬領域データ
cheekAreaData = np.array(data5).reshape(max_row,8)

#画像を入力
#対象とする画像までのパスを入力してください
#img = cv2.imread(r"D:\ryouta\Program\1014_Pre_experiment\honda\ex\landmark\result\re_img_0_.jpg")
#画像ファイルを取得
files = []
#対象とする画像が入っているファイルまでのパスを入力してください。
for f_name in glob.iglob(r"C:\Users\horii\Desktop\sagyou\program\elderly_autocheek\A\10\practice2\after\re\*.jpg"):
    files.append(f_name)
files = natsorted(files)



img_num = len(files)
for i in range(img_num):
    img = cv2.imread(files[i])


    #相対座標に用いる鼻2点の座標を赤く描写
    nose_right_point = [noseAreaData[i][8], noseAreaData[i][9]]
    img = cv2.circle(img, (nose_right_point[0],nose_right_point[1]), 2, (0,0,255), -1)
    nose_left_point = [noseAreaData[i][16], noseAreaData[i][17]]
    img = cv2.circle(img, (nose_left_point[0],nose_left_point[1]), 2, (0,0,255), -1)


    
 
    right_x = cheekAreaData[i][0]
    right_y = cheekAreaData[i][1]
    left_x = cheekAreaData[i][4]
    left_y = cheekAreaData[i][5]
    img = cv2.rectangle(img, (right_x,right_y), (cheekAreaData[i][2],cheekAreaData[i][3]), (255,0,0), 1)
    img = cv2.rectangle(img, (left_x ,left_y), (cheekAreaData[i][6],cheekAreaData[i][7]), (255,0,0), 1)

    """
    #鼻の基準点から頬領域まで直線を描画
    img = cv2.line(img, (nose_right_point[0],nose_right_point[1]), (right_x+30,right_y+15), (255,0,0), 1)
    img = cv2.line(img, (nose_left_point[0],nose_left_point[1]), (left_x,left_y+15), (255,0,0), 1)
    """

    #相対座標を取得
    right_relative_coordinates = [right_x - nose_right_point[0], right_y - nose_right_point[1]]
    left_relative_coordinates = [left_x - nose_left_point[0], left_y - nose_left_point[1]]

    right_relative_coordinates = np.array(right_relative_coordinates)
    left_relative_coordinates = np.array(left_relative_coordinates)

        
    #処理後の画像を出力
    #画像を出力するファイルまでのパスを入力してください
    #cv2.imwrite(r"D:\ryouta\Program\1014_Pre_experiment\honda\ex\landmark\cheek_area.jpg",img)
    cv2.imwrite(r"C:\Users\horii\Desktop\sagyou\program\elderly_autocheek\A\10\practice2\after\cheek\cheek_area" + str(i) + ".jpg",img)




    #相対座標データをxlsxファイルに出力
    ws.merge_cells("B1:C1")
    ws.cell(row=1, column=2, value="頬領域(右頬)")
    ws.merge_cells("D1:E1")
    ws.cell(row=1, column=4, value="頬領域(左頬)")
    ws.cell(row=2, column=1, value="B")
    p = 2
    for val1 in right_relative_coordinates:
        ws.cell(row=i+2, column=p, value=val1)
        p += 1
    for val2 in left_relative_coordinates:
        ws.cell(row=i+2, column=p, value=val2)
        p += 1


#Excelファイルを保存
#Excelファイルを出力するファイルまでのパスを入力してください
#wb.save(r"D:\ryouta\Program\1014_Pre_experiment\honda\ex\landmark\outfile\RelativeCoordinatesData.xlsx")
wb.save(r"C:\Users\horii\Desktop\sagyou\program\elderly_autocheek\A\10\practice2\after\excel\RelativeCoordinatesData.xlsx")
