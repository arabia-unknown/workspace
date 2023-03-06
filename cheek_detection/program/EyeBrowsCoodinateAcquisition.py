import dlib
import cv2
from imutils import face_utils
import openpyxl
import glob
import numpy as np
from natsort import natsorted


detector = dlib.get_frontal_face_detector()

predictor = dlib.shape_predictor("shape_predictor_68_face_landmarks.dat")

flag1 = 0
r = 1
s = 1
y = 1
u = 1
v = 1

wb2 = openpyxl.Workbook()
wb3 = openpyxl.Workbook()
wb4 = openpyxl.Workbook()
wb5 = openpyxl.Workbook()
wb6 = openpyxl.Workbook()

ws2 = wb2.active
ws3 = wb3.active
ws4 = wb4.active
ws5 = wb5.active
ws6 = wb6.active



files = []

for f_name in glob.iglob(r"C:\Users\dot\Desktop\che\program\face\*.jpg"):
    files.append(f_name)
files = natsorted(files)

img_num = len(files)
for i in range(img_num):
    img = cv2.imread(files[i])

    rects = detector(img)
    flag2 = 0

    eyebrows_point = []
    eye_point = []
    contours_point = []

    for rect in rects:
        if (flag2 == 0):

            shape = predictor(img, rect)

            shape = face_utils.shape_to_np(shape)
    

            eye_point = shape[36:48]

            contours_point = shape[0:17]

            eyebrows_point = shape[17:27]

            nose_point = shape[27:36]

            mouth_point = shape[48:68]

            t = 0
            i = 0
            j = 0
            k = 0
            l = 0
            m = 0
            


            while t <= 11:
                img = cv2.circle(img, (eye_point[t][0],eye_point[t][1]), 2, (0,255,0), -1)
                t += 1

            while j <= 9:
                img = cv2.circle(img, (eyebrows_point[j][0],eyebrows_point[j][1]), 2, (0,255,0), -1)
                j += 1

            while k <= 16:
                img = cv2.circle(img, (contours_point[k][0],contours_point[k][1]), 2, (0,255,0), -1)
                k += 1

            while l <= 8:
                img = cv2.circle(img, (nose_point[l][0],nose_point[l][1]), 2, (0,255,0), -1)
                l += 1

            while m <= 19:
                img = cv2.circle(img, (mouth_point[m][0],mouth_point[m][1]), 2, (0,255,0), -1)
                m += 1
            flag2 += 1

            img = cv2.circle(img, (nose_point[2][0],nose_point[2][1]), 2, (0,0,255), -1)
            img = cv2.circle(img, (nose_point[6][0],nose_point[6][1]), 2, (0,0,255), -1)
            img = cv2.circle(img, (contours_point[4][0],contours_point[4][1]), 2, (0,0,255), -1)
            img = cv2.circle(img, (contours_point[12][0],contours_point[12][1]), 2, (0,0,255), -1)
            img = cv2.circle(img, (mouth_point[0][0],mouth_point[0][1]), 2, (0,0,255), -1)
            img = cv2.circle(img, (mouth_point[6][0],mouth_point[6][1]), 2, (0,0,255), -1)
            a=img[shape[29][1]:shape[33][1], shape[54][0]:shape[12][0]] #right cheeks
            b=img[shape[29][1]:shape[33][1], shape[4][0]:shape[48][0]] #left cheek
    
    # img=cv2.rectangle(img,(shape[54][0],shape[29][1]),(shape[12][0],shape[33][1]),(255,0,0), 1)
    # img=cv2.rectangle(img,(shape[4][0],shape[29][1]),(shape[48][0],shape[33][1]),(255,0,0), 1)


    
    r1cheek = [shape[54][0],shape[29][1]]
    r2cheek = [shape[12][0],shape[33][1]]
    l1cheek = [shape[4][0],shape[29][1]]
    l2cheek = [shape[48][0],shape[33][1]]

    centerRx =  (np.array(shape[54][0]) + np.array(shape[12][0]))//2
    centerRy =  (np.array(shape[29][1]) + np.array(shape[33][1]))//2
    centerLx =  (np.array(shape[4][0]) + np.array(shape[48][0]))//2
    centerLy =  (np.array(shape[29][1]) + np.array(shape[33][1]))//2

    img=cv2.rectangle(img,(centerRx - 15,centerRy - 15),(centerRx + 15,centerRy + 15),(0,255,0), 1)
    img=cv2.rectangle(img,(centerLx - 15,centerLy - 15),(centerLx + 15,centerLy + 15),(0,255,0), 1)
    img = cv2.circle(img, (centerRx,centerRy), 2, (0,0,255), -1)
    img = cv2.circle(img, (centerLx,centerLy), 2, (0,0,255), -1)

    r1cheek = np.array(r1cheek)
    r2cheek = np.array(r2cheek)
    l1cheek = np.array(l1cheek)
    l2cheek = np.array(l2cheek)
    

    cv2.imwrite(r"C:\Users\dot\Desktop\che\program\dlib\l" +str(flag1) + ".jpg",img)
    flag1 += 1
    
    

    count = 1
    for val in eyebrows_point:        
        ws3.cell(row=s, column=count, value=val[0])
        count += 1
        ws3.cell(row=s, column=count, value=val[1])
        count += 1
    s += 1
    

    count = 1
    for val in contours_point:        
        ws4.cell(row=y, column=count, value=val[0])
        count += 1
        ws4.cell(row=y, column=count, value=val[1])
        count += 1
    y += 1
    

    count = 1
    for val in nose_point:        
        ws5.cell(row=u, column=count, value=val[0])
        count += 1
        ws5.cell(row=u, column=count, value=val[1])
        count += 1
    u += 1


    # ws6.merge_cells("A1:B1")
    # ws6.cell(row=1, column=1, value="頬領域(右頬矩形右上x,y)")
    # ws6.merge_cells("C1:D1")
    # ws6.cell(row=1, column=3, value="頬領域(右頬矩形左上x,y)")
    # ws6.merge_cells("E1:F1")
    # ws6.cell(row=1, column=5, value="頬領域(左頬矩形右上x,y)")
    # ws6.merge_cells("G1:H1")
    # ws6.cell(row=1, column=7, value="頬領域(左頬矩形左上x,y)")
    count = 1
    for val1 in r1cheek:
        ws6.cell(row=v, column=count, value=val1)
        count += 1
    for val2 in r2cheek:
        ws6.cell(row=v, column=count, value=val2)
        count += 1    
    for val3 in l1cheek:
        ws6.cell(row=v, column=count, value=val3)
        count += 1
    for val4 in l2cheek:
        ws6.cell(row=v, column=count, value=val4)
        count += 1
    v += 1


wb3.save(r"C:\Users\dot\Desktop\che\program\excel\Eyebrows.xlsx")
wb4.save(r"C:\Users\dot\Desktop\che\program\excel\Contours.xlsx")
wb5.save(r"C:\Users\dot\Desktop\che\program\excel\nose.xlsx")
wb6.save(r"C:\Users\dot\Desktop\che\program\excel\cheek.xlsx")
